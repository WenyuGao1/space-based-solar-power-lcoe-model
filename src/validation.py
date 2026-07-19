"""Validation checks for model inputs and generated outputs."""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
import math
import re
from urllib.parse import urlparse

from .model import calculate_lcoe
from .parameters import Parameter


REQUIRED_PARAMETERS = {
    "delivered_capacity_mw",
    "end_to_end_efficiency",
    "capacity_factor",
    "system_lifetime_years",
    "wacc",
    "specific_mass_kg_per_kw_space_power",
    "space_hardware_cost_gbp_per_w_space",
    "wireless_power_transmission_cost_gbp_per_w_space",
    "launch_cost_gbp_per_kg",
    "orbit_transfer_cost_gbp_per_kg",
    "in_orbit_assembly_cost_gbp_per_kg",
    "rectenna_cost_gbp_per_w_delivered",
    "grid_connection_cost_gbp_per_kw_delivered",
    "programme_margin_pct",
    "replacement_refurbishment_pct_capex_per_year",
    "fixed_opex_pct_capex_per_year",
    "variable_opex_gbp_per_mwh",
}

EXPECTED_CSV_SCHEMAS = {
    "assumptions.csv": ["assumption_id", "area", "assumption", "source_type", "source_id", "limitation"],
    "external_sbsp_studies.csv": [
        "source_id",
        "study_label",
        "analysis_type",
        "reported_cost_context",
        "what_it_answers",
        "relationship_to_this_project",
        "comparability_limit",
    ],
    "parameter_evidence.csv": [
        "parameter_or_claim",
        "parameter_or_claim_zh",
        "evidence_role",
        "evidence_role_zh",
        "source_id",
        "locator",
        "supported_evidence",
        "supported_evidence_zh",
        "numeric_context",
        "numeric_context_zh",
        "limitations",
        "limitations_zh",
    ],
    "sbsp_parameters.csv": [
        "parameter",
        "display_name",
        "unit",
        "reference_value",
        "min_value",
        "max_value",
        "improvement_direction",
        "source_type",
        "source_id",
        "notes",
    ],
    "source_registry.csv": [
        "source_id",
        "title",
        "organization",
        "document_year",
        "publication_date",
        "document_revision_date",
        "landing_page_last_updated",
        "accessed_date",
        "url",
        "role",
        "notes",
    ],
    "uk_generation_costs.csv": [
        "technology",
        "category",
        "value_low_gbp_per_mwh",
        "value_mid_gbp_per_mwh",
        "value_high_gbp_per_mwh",
        "price_basis",
        "source_id",
        "value_type",
        "include_in_bands",
        "notes",
    ],
    "uk_system_adjusted_costs.csv": [
        "technology",
        "benchmark_reference",
        "value_low_gbp_per_mwh",
        "value_mid_gbp_per_mwh",
        "value_high_gbp_per_mwh",
        "price_basis",
        "source_id",
        "components",
        "notes",
    ],
}

CSV_PRIMARY_KEYS = {
    "assumptions.csv": "assumption_id",
    "external_sbsp_studies.csv": "source_id",
    "sbsp_parameters.csv": "parameter",
    "source_registry.csv": "source_id",
}


def validate_parameters(params: dict[str, Parameter]) -> list[str]:
    errors: list[str] = []
    missing = REQUIRED_PARAMETERS.difference(params)
    if missing:
        errors.append(f"Missing required parameters: {', '.join(sorted(missing))}")
    for parameter in params.values():
        values = (parameter.min_value, parameter.reference_value, parameter.max_value)
        if not all(math.isfinite(value) for value in values):
            errors.append(f"{parameter.name} contains a non-finite value.")
        if parameter.min_value > parameter.max_value:
            errors.append(f"{parameter.name} has min greater than max.")
        if not (parameter.min_value <= parameter.reference_value <= parameter.max_value):
            errors.append(f"{parameter.name} reference value is outside the explored range.")
        if parameter.improvement_direction not in {"lower", "higher"}:
            errors.append(f"{parameter.name} has invalid improvement direction.")
        if parameter.name in {"end_to_end_efficiency", "capacity_factor"}:
            if parameter.min_value <= 0.0 or parameter.max_value > 1.0:
                errors.append(f"{parameter.name} explored range must lie in (0, 1].")
        elif parameter.name == "wacc":
            if parameter.min_value < 0.0 or parameter.max_value >= 1.0:
                errors.append("wacc explored range must lie in [0, 1).")
        elif parameter.name in {
            "programme_margin_pct",
            "replacement_refurbishment_pct_capex_per_year",
            "fixed_opex_pct_capex_per_year",
        }:
            if parameter.min_value < 0.0 or parameter.max_value > 1.0:
                errors.append(f"{parameter.name} explored range must lie in [0, 1].")
        elif parameter.name in {
            "delivered_capacity_mw",
            "system_lifetime_years",
            "specific_mass_kg_per_kw_space_power",
        }:
            if parameter.min_value <= 0.0:
                errors.append(f"{parameter.name} explored range must be strictly positive.")
        elif parameter.min_value < 0.0:
            errors.append(f"{parameter.name} explored range cannot be negative.")
    return errors


def validate_reference_case(reference: dict[str, float]) -> list[str]:
    errors: list[str] = []
    if not 0.0 < reference["end_to_end_efficiency"] <= 1.0:
        errors.append("End-to-end efficiency must be in (0, 1].")
    if not 0.0 < reference["capacity_factor"] <= 1.0:
        errors.append("Capacity factor must be in (0, 1].")
    if reference["system_lifetime_years"] <= 0:
        errors.append("System lifetime must be positive.")
    try:
        result = calculate_lcoe(reference)
        if not math.isfinite(result.lcoe_gbp_per_mwh) or result.lcoe_gbp_per_mwh <= 0:
            errors.append("Reference LCOE is not positive.")
        if not math.isclose(
            result.initial_capex_gbp,
            result.pre_margin_capex_gbp + result.programme_margin_gbp,
            rel_tol=1e-12,
        ):
            errors.append("Initial CAPEX does not reconcile to pre-margin CAPEX plus programme margin.")
        annual_cost_check = (
            result.annualized_capex_gbp
            + result.fixed_opex_gbp_per_year
            + result.replacement_refurbishment_gbp_per_year
            + result.variable_opex_gbp_per_year
        )
        if not math.isclose(result.annual_total_cost_gbp, annual_cost_check, rel_tol=1e-12):
            errors.append("Annual total cost does not reconcile to its components.")
    except Exception as exc:  # pragma: no cover - defensive validation
        errors.append(f"Reference case failed: {exc}")
    return errors


def validate_analysis_results(
    reference: dict[str, float],
    params: dict[str, Parameter],
    sensitivity_parameter_names: list[str],
    thresholds: list[dict[str, object]],
    launch_frontier: list[dict[str, object]],
    combined_frontier: list[dict[str, object]],
    alternative_frontiers: list[dict[str, object]],
    tolerance: float = 1e-7,
) -> list[str]:
    """Validate monotonicity, root residuals, boundary failures and frontier states."""

    errors: list[str] = []
    for name in sensitivity_parameter_names:
        parameter = params[name]
        lcoes: list[float] = []
        for index in range(41):
            value = parameter.min_value + (parameter.max_value - parameter.min_value) * index / 40.0
            case = dict(reference)
            case[name] = value
            lcoes.append(calculate_lcoe(case).lcoe_gbp_per_mwh)
        differences = [right - left for left, right in zip(lcoes, lcoes[1:])]
        if parameter.improvement_direction == "lower" and any(delta < -tolerance for delta in differences):
            errors.append(f"One-way LCOE is not monotonic increasing for lower-is-better input: {name}")
        if parameter.improvement_direction == "higher" and any(delta > tolerance for delta in differences):
            errors.append(f"One-way LCOE is not monotonic decreasing for higher-is-better input: {name}")

    for row in thresholds:
        name = str(row["parameter"])
        target = float(row["target_lcoe_gbp_per_mwh"])
        threshold = row.get("threshold_value")
        parameter = params[name]
        if threshold in (None, ""):
            favourable = parameter.min_value if parameter.improvement_direction == "lower" else parameter.max_value
            case = dict(reference)
            case[name] = favourable
            if calculate_lcoe(case).lcoe_gbp_per_mwh <= target + tolerance:
                errors.append(f"One-way target marked unreachable although favourable bound is feasible: {name} / {target}")
            continue
        value = float(threshold)
        if not parameter.min_value <= value <= parameter.max_value:
            errors.append(f"One-way threshold is outside its parameter bounds: {name} / {target}")
            continue
        case = dict(reference)
        case[name] = value
        solved_lcoe = calculate_lcoe(case).lcoe_gbp_per_mwh
        boundary_solution = math.isclose(value, parameter.min_value, abs_tol=1e-12) or math.isclose(
            value, parameter.max_value, abs_tol=1e-12
        )
        if boundary_solution:
            if solved_lcoe > target + tolerance:
                errors.append(f"One-way boundary solution does not meet target: {name} / {target}")
        elif not math.isclose(solved_lcoe, target, abs_tol=tolerance):
            errors.append(f"One-way root residual exceeds tolerance: {name} / {target}")

    for row in launch_frontier:
        target = float(row["target_lcoe_gbp_per_mwh"])
        efficiency = float(row["end_to_end_efficiency"])
        threshold = row.get("max_launch_cost_gbp_per_kg")
        case = dict(reference)
        case["end_to_end_efficiency"] = efficiency
        if threshold in (None, ""):
            case["launch_cost_gbp_per_kg"] = params["launch_cost_gbp_per_kg"].min_value
            if calculate_lcoe(case).lcoe_gbp_per_mwh <= target + tolerance:
                errors.append(f"Launch frontier marked unreachable although lower bound is feasible: {efficiency} / {target}")
        else:
            case["launch_cost_gbp_per_kg"] = float(threshold)
            if not math.isclose(calculate_lcoe(case).lcoe_gbp_per_mwh, target, abs_tol=tolerance):
                errors.append(f"Launch-frontier root residual exceeds tolerance: {efficiency} / {target}")

    def validate_coupled_rows(rows: list[dict[str, object]], label: str) -> None:
        groups: dict[str, list[dict[str, object]]] = {}
        for row in rows:
            group = str(row.get("pathway", "combined"))
            groups.setdefault(group, []).append(row)
            status = str(row.get("status", ""))
            progress = row.get("progress_fraction")
            target = float(row["target_lcoe_gbp_per_mwh"])
            if status in {"feasible", "already feasible at baseline", "already feasible at slice baseline"}:
                if progress in (None, ""):
                    errors.append(f"{label} row has no progress value: {group} / {target}")
                    continue
                progress_value = float(progress)
                if not 0.0 <= progress_value <= 1.0:
                    errors.append(f"{label} progress is outside [0,1]: {group} / {target}")
                case = dict(reference)
                for name in params:
                    if name in row:
                        case[name] = float(row[name])
                solved_lcoe = calculate_lcoe(case).lcoe_gbp_per_mwh
                if status == "feasible" and not math.isclose(solved_lcoe, target, abs_tol=tolerance):
                    errors.append(f"{label} root residual exceeds tolerance: {group} / {target}")
                if status.startswith("already feasible"):
                    if not math.isclose(progress_value, 0.0, abs_tol=1e-12) or solved_lcoe > target + tolerance:
                        errors.append(f"{label} baseline-feasible status is inconsistent: {group} / {target}")
            elif status != "not feasible within explored combined range":
                errors.append(f"{label} has unknown status {status!r}: {group} / {target}")

        for group, group_rows in groups.items():
            ordered = sorted(group_rows, key=lambda row: float(row["target_lcoe_gbp_per_mwh"]), reverse=True)
            progress_values = [
                float(row["progress_fraction"])
                for row in ordered
                if row.get("progress_fraction") not in (None, "")
            ]
            if any(right + tolerance < left for left, right in zip(progress_values, progress_values[1:])):
                errors.append(f"{label} progress is not monotonic as targets tighten: {group}")

    validate_coupled_rows(combined_frontier, "Combined frontier")
    validate_coupled_rows(alternative_frontiers, "Alternative frontier")
    return errors


def validate_raw_workbooks(paths: list[str | Path]) -> list[str]:
    """Reject failed downloads and workbooks that cannot be parsed as XLSX."""
    errors: list[str] = []
    for path_like in paths:
        path = Path(path_like)
        if not path.exists():
            errors.append(f"Missing raw workbook: {path}")
            continue
        with path.open("rb") as handle:
            signature = handle.read(4)
        if signature != b"PK\x03\x04":
            errors.append(f"Raw workbook is not a valid XLSX ZIP container: {path}")
            continue
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            if not workbook.sheetnames:
                errors.append(f"Raw workbook contains no worksheets: {path}")
            workbook.close()
        except Exception as exc:
            errors.append(f"Raw workbook cannot be parsed as XLSX: {path} ({exc})")
    return errors


def validate_csv_structure(paths: list[str | Path]) -> list[str]:
    """Validate exact schemas, record widths, keys, enums and numeric ordering."""

    errors: list[str] = []
    for path_like in paths:
        path = Path(path_like)
        if not path.exists():
            errors.append(f"Missing CSV input: {path}")
            continue
        with path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            headers = reader.fieldnames or []
            if not headers:
                errors.append(f"CSV has no header: {path}")
                continue
            expected_schema = EXPECTED_CSV_SCHEMAS.get(path.name)
            if expected_schema is not None and headers != expected_schema:
                errors.append(
                    f"CSV schema mismatch for {path.name}: expected {expected_schema}, found {headers}"
                )
            if len(headers) != len(set(headers)):
                errors.append(f"CSV has duplicate header names: {path}")
            if any(not str(header).strip() for header in headers):
                errors.append(f"CSV has a blank header name: {path}")
            primary_key = CSV_PRIMARY_KEYS.get(path.name)
            seen_keys: set[str] = set()
            for line_number, row in enumerate(reader, start=2):
                if None in row:
                    errors.append(f"CSV row has unquoted extra fields: {path}:{line_number}")
                missing = [header for header in headers if row.get(header) is None]
                if missing:
                    errors.append(
                        f"CSV row is shorter than its header: {path}:{line_number} "
                        f"(missing {', '.join(missing)})"
                    )
                if primary_key:
                    key_value = (row.get(primary_key) or "").strip()
                    if not key_value:
                        errors.append(f"Blank primary key {primary_key} at {path}:{line_number}")
                    elif key_value in seen_keys:
                        errors.append(f"Duplicate primary key {key_value!r} at {path}:{line_number}")
                    seen_keys.add(key_value)

                source_id = (row.get("source_id") or "").strip()
                if "source_id" in headers and not source_id:
                    errors.append(f"Blank source_id at {path}:{line_number}")

                if path.name == "sbsp_parameters.csv":
                    try:
                        low = float(row["min_value"])
                        reference = float(row["reference_value"])
                        high = float(row["max_value"])
                        if not all(math.isfinite(value) for value in (low, reference, high)):
                            errors.append(f"Non-finite parameter value at {path}:{line_number}")
                        if not low <= reference <= high:
                            errors.append(f"Parameter range ordering fails at {path}:{line_number}")
                    except (TypeError, ValueError):
                        errors.append(f"Invalid numeric parameter value at {path}:{line_number}")
                    if row.get("improvement_direction") not in {"lower", "higher"}:
                        errors.append(f"Invalid improvement_direction at {path}:{line_number}")
                    if row.get("source_type") not in {"exploratory", "sourced", "derived"}:
                        errors.append(f"Invalid source_type at {path}:{line_number}")

                if path.name in {"uk_generation_costs.csv", "uk_system_adjusted_costs.csv"}:
                    numeric_values: list[float] = []
                    for field in (
                        "value_low_gbp_per_mwh",
                        "value_mid_gbp_per_mwh",
                        "value_high_gbp_per_mwh",
                    ):
                        value = (row.get(field) or "").strip()
                        if value:
                            try:
                                parsed = float(value)
                                if not math.isfinite(parsed):
                                    raise ValueError
                                numeric_values.append(parsed)
                            except ValueError:
                                errors.append(f"Invalid benchmark number at {path}:{line_number} field {field}")
                    if numeric_values and len(numeric_values) != 3:
                        errors.append(f"Partially populated benchmark range at {path}:{line_number}")
                    if len(numeric_values) == 3 and numeric_values != sorted(numeric_values):
                        errors.append(f"Benchmark low/mid/high ordering fails at {path}:{line_number}")
                if path.name == "uk_generation_costs.csv" and row.get("include_in_bands") not in {"true", "false"}:
                    errors.append(f"Invalid include_in_bands boolean at {path}:{line_number}")
    return errors


def validate_source_integrity(
    source_registry_path: str | Path,
    linked_csv_paths: list[str | Path],
) -> list[str]:
    """Validate source metadata and foreign keys used by analytical inputs."""

    errors: list[str] = []
    registry_path = Path(source_registry_path)
    with registry_path.open(newline="", encoding="utf-8") as handle:
        registry_rows = list(csv.DictReader(handle))

    source_ids = [row.get("source_id", "").strip() for row in registry_rows]
    if len(source_ids) != len(set(source_ids)):
        errors.append("Source registry contains duplicate source_id values.")
    if any(not source_id for source_id in source_ids):
        errors.append("Source registry contains a blank source_id.")
    known_sources = set(source_ids)

    for row in registry_rows:
        source_id = row.get("source_id", "<blank>")
        for field in (
            "publication_date",
            "document_revision_date",
            "landing_page_last_updated",
            "accessed_date",
        ):
            value = row.get(field, "").strip()
            if not value:
                if field == "accessed_date":
                    errors.append(f"{source_id} has no accessed_date.")
                continue
            try:
                date.fromisoformat(value)
            except ValueError:
                errors.append(f"{source_id} has invalid {field}: {value!r}")
        document_year = row.get("document_year", "").strip()
        if document_year != "live" and not (len(document_year) == 4 and document_year.isdigit()):
            errors.append(f"{source_id} has invalid document_year: {document_year!r}")
        url = row.get("url", "").strip()
        if url.startswith("http"):
            parsed = urlparse(url)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                errors.append(f"{source_id} has an invalid URL: {url}")
        elif url.startswith("report/"):
            project_path = registry_path.parent.parent / url
            if not project_path.exists():
                errors.append(f"{source_id} references a missing project file: {url}")
        else:
            errors.append(f"{source_id} has a non-resolvable URL or project path: {url}")

    for path_like in linked_csv_paths:
        path = Path(path_like)
        with path.open(newline="", encoding="utf-8") as handle:
            for line_number, row in enumerate(csv.DictReader(handle), start=2):
                source_id = row.get("source_id", "").strip()
                if not source_id:
                    errors.append(f"Blank source_id at {path}:{line_number}")
                elif source_id not in known_sources:
                    errors.append(f"Unknown source_id {source_id!r} at {path}:{line_number}")
                if path.name == "sbsp_parameters.csv" and source_id != "ASSUMPTION_THIS_STUDY":
                    errors.append(
                        f"Exact model input is not classified as study-authored at {path}:{line_number}"
                    )
    return errors


def validate_official_generation_extract(
    workbook_path: str | Path,
    generation_csv_path: str | Path,
) -> list[str]:
    """Reconcile selected DESNZ 2025 workbook totals to the benchmark CSV."""

    errors: list[str] = []
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"Could not load official generation-cost workbook: {exc}"]

    expected_cells = {
        "Additional Estimates 2030": {
            "A2": "All costs in 2024 real GBP prices",
            "C15": 60,
            "D15": 58,
            "E15": 103,
            "F15": 153,
            "I15": 111,
            "AC15": 181,
            "AD15": 105,
        },
        "Additional Estimates 2035": {
            "A2": "All costs in 2024 real GBP prices",
            "C15": 55,
            "D15": 57,
            "E15": 98,
            "F15": 125,
            "I15": 113,
            "AC15": 181,
            "AD15": 104,
        },
    }
    for sheet_name, cells in expected_cells.items():
        if sheet_name not in workbook.sheetnames:
            errors.append(f"Official workbook is missing sheet: {sheet_name}")
            continue
        sheet = workbook[sheet_name]
        for address, expected in cells.items():
            actual = sheet[address].value
            if actual != expected:
                errors.append(
                    f"Official workbook mismatch at {sheet_name}!{address}: "
                    f"expected {expected!r}, found {actual!r}"
                )
    workbook.close()

    expected_ranges = {
        "Large-scale solar": (55.0, 57.5, 60.0),
        "Onshore wind": (57.0, 57.5, 58.0),
        "Fixed offshore wind": (98.0, 100.5, 103.0),
        "Floating offshore wind": (125.0, 139.0, 153.0),
        "Gas CCGT high load factor": (111.0, 112.0, 113.0),
        "Gas with CCUS high load factor": (104.0, 104.5, 105.0),
        "Gas with CCUS mid load factor": (181.0, 181.0, 181.0),
    }
    with Path(generation_csv_path).open(newline="", encoding="utf-8") as handle:
        rows = {row["technology"]: row for row in csv.DictReader(handle)}
    for technology, expected in expected_ranges.items():
        if technology not in rows:
            errors.append(f"Generation benchmark extract is missing: {technology}")
            continue
        row = rows[technology]
        actual = tuple(
            float(row[field])
            for field in (
                "value_low_gbp_per_mwh",
                "value_mid_gbp_per_mwh",
                "value_high_gbp_per_mwh",
            )
        )
        if actual != expected:
            errors.append(f"Generation benchmark mismatch for {technology}: {actual} != {expected}")
    return errors


def validate_outputs(paths: list[str | Path]) -> list[str]:
    errors: list[str] = []
    for path_like in paths:
        path = Path(path_like)
        if not path.exists():
            errors.append(f"Missing output: {path}")
        elif path.is_file() and path.stat().st_size == 0:
            errors.append(f"Output is empty: {path}")
    return errors


def validate_generated_csv_outputs(
    specifications: list[dict[str, object]],
) -> list[str]:
    """Check generated CSV row counts, required columns and compound keys."""

    errors: list[str] = []
    for specification in specifications:
        path = Path(str(specification["path"]))
        if not path.exists():
            errors.append(f"Missing generated CSV: {path}")
            continue
        with path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
            headers = reader.fieldnames or []
        expected_rows = int(specification["row_count"])
        if len(rows) != expected_rows:
            errors.append(f"Generated CSV row count mismatch for {path}: {len(rows)} != {expected_rows}")
        required_columns = [str(value) for value in specification.get("required_columns", [])]
        missing_columns = [column for column in required_columns if column not in headers]
        if missing_columns:
            errors.append(f"Generated CSV {path} is missing columns: {', '.join(missing_columns)}")
        key_columns = [str(value) for value in specification.get("key_columns", [])]
        if key_columns and not missing_columns:
            keys = [tuple(row.get(column, "") for column in key_columns) for row in rows]
            if len(keys) != len(set(keys)):
                errors.append(f"Generated CSV {path} has duplicate key values for {key_columns}")
    return errors


def validate_png_images(paths: list[str | Path], minimum_width: int = 800, minimum_height: int = 500) -> list[str]:
    """Open every PNG and reject truncated or implausibly small analytical figures."""

    errors: list[str] = []
    try:
        from PIL import Image
    except Exception as exc:  # pragma: no cover - dependency validation
        return [f"Pillow unavailable for PNG validation: {exc}"]
    for path_like in paths:
        path = Path(path_like)
        if not path.exists():
            errors.append(f"Missing PNG image: {path}")
            continue
        try:
            with Image.open(path) as image:
                image.verify()
            with Image.open(path) as image:
                if image.format != "PNG":
                    errors.append(f"Analytical image is not PNG: {path}")
                if image.width < minimum_width or image.height < minimum_height:
                    errors.append(
                        f"Analytical image is too small: {path} ({image.width}x{image.height})"
                    )
        except Exception as exc:
            errors.append(f"PNG cannot be decoded: {path} ({exc})")
    return errors


def validate_markdown_images(report_paths: list[str | Path], expected_count: int = 7) -> list[str]:
    """Resolve every Markdown image and enforce the intended main-figure count."""

    errors: list[str] = []
    pattern = re.compile(r"!\[[^\]]*\]\(([^)]+)\)")
    for path_like in report_paths:
        path = Path(path_like)
        if not path.exists():
            errors.append(f"Missing Markdown report: {path}")
            continue
        links = pattern.findall(path.read_text(encoding="utf-8"))
        if len(links) != expected_count:
            errors.append(f"Markdown report {path.name} has {len(links)} images; expected {expected_count}")
        if len(links) != len(set(links)):
            errors.append(f"Markdown report {path.name} repeats a main image link.")
        for link in links:
            target = (path.parent / link).resolve()
            if not target.exists():
                errors.append(f"Markdown report {path.name} references a missing image: {link}")
    return errors


def validate_report_content(
    report_paths: list[str | Path],
    required_tokens: list[str],
) -> list[str]:
    """Confirm that both generated language reports carry the shared key facts."""

    errors: list[str] = []
    for path_like in report_paths:
        path = Path(path_like)
        if not path.exists():
            errors.append(f"Missing report for content validation: {path}")
            continue
        text = path.read_text(encoding="utf-8")
        for token in required_tokens:
            if token not in text:
                errors.append(f"Generated report {path.name} is missing required token: {token!r}")
    return errors


def validate_pdf_documents(
    paths: list[str | Path],
    minimum_pages: int = 8,
    minimum_images: int = 7,
    required_text_tokens: list[str] | None = None,
) -> list[str]:
    """Check PDF structure, page count and extractable text."""

    errors: list[str] = []
    try:
        from pypdf import PdfReader
    except Exception as exc:  # pragma: no cover - dependency validation
        return [f"pypdf unavailable for PDF validation: {exc}"]

    for path_like in paths:
        path = Path(path_like)
        if not path.exists():
            errors.append(f"Missing PDF document: {path}")
            continue
        try:
            reader = PdfReader(path)
            if len(reader.pages) < minimum_pages:
                errors.append(
                    f"PDF has fewer than {minimum_pages} pages: {path} ({len(reader.pages)})"
                )
            extracted = "".join((page.extract_text() or "") for page in reader.pages)
            if len(extracted.strip()) < 500:
                errors.append(f"PDF has insufficient extractable text: {path}")
            for token in required_text_tokens or []:
                if token not in extracted:
                    errors.append(f"PDF {path.name} is missing required text token: {token!r}")
            image_count = sum(len(page.images) for page in reader.pages)
            if image_count < minimum_images:
                errors.append(
                    f"PDF has fewer than {minimum_images} embedded images: {path} ({image_count})"
                )
            metadata = reader.metadata or {}
            if metadata.get("/Author") != "Wenyu Gao":
                errors.append(f"PDF author metadata is not Wenyu Gao: {path}")
            if "Cost-Condition" not in str(metadata.get("/Title", "")) and "成本条件" not in str(metadata.get("/Title", "")):
                errors.append(f"PDF title metadata is not the v1.2 cost-condition title: {path}")
        except Exception as exc:
            errors.append(f"PDF cannot be parsed: {path} ({exc})")
    return errors
